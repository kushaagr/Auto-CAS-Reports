import config
import departmentdata as college
import recommendations as reco
from dirutility import clean_up

import heapq
import pathlib
import random
import string
import sqlite3
import os
import collections as coll
import openpyxl as op
from fpdf import FPDF
from typing import List


# Module's private variable 
__VOWELS      = ('a', 'e', 'i', 'o', 'u')
__CONSONANTS  = tuple(set(string.ascii_lowercase) - set(__VOWELS))

__DEF_FSIZE   = 12
__PG_WIDTH    = 210 #mm
__PG_HEIGHT   = 297 #mm
# __GRAPHS_DIR  = './temp'
__GRAPHS_DIR  = config.TEMPDIR
__IMGDIR  = './images'
__IMGS    = ('report-newpage-1.png', 'report-newpage-2.png', 'report-newpage-3.png',
        'report-newpage-4.png', 'report-newpage-5.png', 'report-newpage-6.png')

TEMPLATE_WITH_CONTACT   = 'template-with-contact-info.png'
ACROCARE_EMAIL          = r'mailto:acrocare@acropolis.in'
AITR_URL = r'https://aitr.ac.in/'
AITRGOOGLE_URL = r'https://www.google.com/search?q=acropolis+institute+of+technology+and+research&source=lmns&bih=568&biw=1366&hl=en&sa=X&ved=2ahUKEwi-7LL_oJD3AhWz_DgGHaamCjEQ_AUoAHoECAEQAA'
debug_setting = config.DEBUG_MODE

# def generate_codename(studata: list, department: str, count: int) -> str:
# def generate_codename(studata: list, department: str) -> str:
def generate_codename(studata: list) -> str:
    """ 
    Return a string in format: 
    <4-char-Department-code><2-char-Year><2-char-Name-initials>
    """

    # deptcode        = college.DEPT_CODE_MAP[department]
    department      = studata[-3]
    deptcode        = college.DEPT_CODE_MAP.get(department, department)
    stuname         = studata[0]
    nameinitials    = ''.join(word[0] for word in stuname.strip().split())
    yearstr         = studata[-2].strip()
    year : int      = 1 # default
    for i in yearstr.split():
        try:
            year    = int(i)
            break
        except ValueError:
            continue

    # nomenclature_str = f"{deptcode}{year}{nameinitials}{count:03}"
    nomenclature_str = f"{deptcode.upper()}{year:02}{nameinitials.upper()}"
    return nomenclature_str


# def generate_codenames_list(data: List[List[str]], whichdepartment: str) -> list:
def generate_codenames_list(data: List[List[str]]) -> list:
    codenames = []
    created_codenames = set()
    print(f'{data=}')
    # for gi, arr in enumerate(data, start=1):
    for gi in range(1, len(data)):
        # codename = ''
        count = 1
        # codename = generate_codename(data[gi], whichdepartment)
        print("index=",gi, end=" ")
        codename = generate_codename(data[gi])
        while (f'{codename}{count:03}' in created_codenames):
            count += 1
        codename = f'{codename}{count:03}'
        codenames.append(codename)
        created_codenames.add(codename)
    return codenames


def generate_codename_e1(name: str, lenlimit: int = None) -> str:
    codename = []
    for c in reversed(name):
        codename.extend( (c.lower(), random.choice(__VOWELS), random.choice(__CONSONANTS)) )
    return ''.join(codename)[:lenlimit].capitalize()


# Unused and deprecated function
def generate_codenames_list_e1(data: list):
    codenames = []
    for i in range(1, len(data)):
        name = data[i][0]
        guardian_name = data[i][2]
        requiredlen = 8
        name = name.split(' ')[0]
        guardian_name = guardian_name.split(' ')[0]
        codename = generate_codename(name, requiredlen)
        if (len(codename) < requiredlen):
            codename = generate_codename(guardian_name+name, requiredlen)
        codenames.append(codename)
    return codenames


def Upload_Summary(dirpath: str, tscores: list, 
                    studata: list, codenames: list, survey_name: str=''):
    header = ('NAME', 
    'CODE NAME',
    # 'DEMOGRAPHICS (AGE, GENDER, INSTITUTE, STREAM, YEAR, MONTHY FAMILY INCOME)',
    # 'CONTACT DETAILS (EMAIL, MOBILE)',
    *studata[0][1:],
    'AN-TSCORE', 'DP-TSCORE', 'SI-TSCORE', 'SA-TSCORE', 'SE-TSCORE', 'IP-TSCORE', 'FP-TSCORE', 'AP-TSCORE', 'CP-TSCORE')
    wb = op.Workbook()
    sheet = wb.active
    sheet.append(header)
    for gi, cname in enumerate(codenames, start=1):
        name    = studata[gi][0]
        sheet.append((name, cname, *studata[gi][1:], *tscores[gi]))

    if (survey_name.strip() != ''):
        survey_name = f" - {survey_name}"
    wb.save(f"{dirpath}/Summary{survey_name}.xlsx")


def plot_tscores(pdfobj, col1: list, col2: list, 
                topproblemareas: list):
    M = 8
    LIGHTBLUE = (224, 235, 255)
    pdfobj.ln(40)
    pdfobj.set_left_margin(M)
    pdfobj.set_right_margin(M)
    pdfobj.set_fill_color(*LIGHTBLUE)
    # pdfobj.set_text_color(0)

    pdfobj.cell(100,9, txt=col1[0], align='C', border=1)
    pdfobj.cell(0,9, txt=str(col2[0]), align='C', border=1)
    pdfobj.ln()

    fill = True
    for category, score in zip(col1[1:], col2[1:]):
        if (category in topproblemareas):
            fill = True
            pdfobj.set_font(style="B")
            # pdfobj.set_fill_color(187, 104, 101)
            # pdfobj.set_fill_color(222, 140, 51)
            # pdfobj.set_fill_color(171, 107, 39)
            # pdfobj.set_fill_color(220, 218, 131)
            # pdfobj.set_fill_color(169, 167, 100)
            pdfobj.set_fill_color(171, 107, 39)
            # pdfobj.set_fill_color(118, 136, 220)
            # pdfobj.set_fill_color(220, 136, 118)

            pdfobj.set_text_color(255)
        else:
            # pdfobj.set_fill_color(224, 235, 255)
            pdfobj.set_font(style="")
            pdfobj.set_fill_color(*LIGHTBLUE)
            pdfobj.set_text_color(0)

        pdfobj.cell(100,8, txt=category, border=1, fill=fill)
        pdfobj.cell(0,8, txt=str(score), align='C', border=1, fill=fill)
        pdfobj.ln()
        fill = not fill
    pdfobj.set_font(style="")
    pdfobj.set_text_color(0)
    pdfobj.set_margins(0, 0, 0)

# Unused and Deprecated function
def plot_problem_areas(pdfobj, col: list, 
                        xpos: int, ypos: int = None):
    width = 60
    pdfobj.set_font_size(20)
    # pdfobj.set_left_margin(7)
    pdfobj.set_line_width(.7)
    pdfobj.set_fill_color(149, 83, 81)
    pdfobj.set_text_color(255)
    pdfobj.ln(ypos)
    pdfobj.set_x(xpos)
    pdfobj.set_font(style='UB')
    pdfobj.cell(width, 10, "Problem Areas", align='C', border='TB', fill=1)
    pdfobj.ln()
    pdfobj.set_x(xpos)
    pdfobj.set_fill_color(201, 112, 109)
    pdfobj.cell(width, 7, fill=1)

    pdfobj.set_font(style='B')
    pdfobj.set_font_size(12)
    for var in col:
        pdfobj.ln()
        pdfobj.set_x(xpos)
        # pdfobj.cell(50, 7, txt=var, align='C', border='TB', fill=1)
        pdfobj.cell(width, 7, txt=var, align='C', fill=1)
    pdfobj.ln()
    pdfobj.set_x(xpos)
    pdfobj.cell(width, 7, border='B', fill=1)
    
    # Revert all styling
    pdfobj.set_font_size(__DEF_FSIZE)
    pdfobj.set_text_color(0)
    pdfobj.set_font(style="")
    # pdfobj.set_fill_color(255)


def Upload_All_Reports(dirpath: str, tscoreslist: list, data: list, 
                        graphs: list, survey_id: int, 
                        whichdepartment: str = None) -> str:

    con1 = sqlite3.connect(config.DB)
    cur = con1.cursor()
    # cur.execute("""SELECT student_name, student_codename FROM tblSurveyReports 
        # WHERE survey_id=?""", (survey_id, ))
    # name_codename_map = {}
    # rows = cur.fetchall()
    # print(cur.fetchall())
    # print(rows)
    # for na, cona in rows:
    #     print(na, cona)
    #     name_codename_map[na] = cona 
        # print(na, cona )
    cur.execute("""DELETE FROM tblSurveyReports WHERE survey_id = ?""",
        (str(survey_id), ))

    created_codenames = set()
    for gi, stu in enumerate(graphs, start=1):
        # tscores       = ['Score', 71,70,67,56,67,23,45,56]
        # problemareas  = ['Anxiety','Depression','Suicidal Ideation','Interpersonalproblems']
        tscores         = ['Score'] + tscoreslist[gi]
        top4tscores     = heapq.nlargest(4, tscoreslist[gi])
        top4count       = coll.Counter(top4tscores)
        problemareas    = []
        for j, t in enumerate(tscores[1:], start=0):
            if top4count.get(t, 0) > 0 and t >= 65:
                top4count[t] -= 1
                problemareas.append(reco.CATEGORIES[1:][j])                                    
            elif (int(t) >= 70):
                problemareas.append(reco.CATEGORIES[1:][j])

        # name = data[gi][0]
        # print("name, codename = ", name, name_codename_map[name])
        # requiredlen = 8
        # if name in name_codename_map:
        #     codename = name_codename_map[name]
        # else:
        #     name = data[gi][0].split(' ')[0]
        #     guardian_name = data[gi][2].split(' ')[0]
        #     codename = generate_codename(name, requiredlen)
        #     if (len(codename) < requiredlen):
        #         codename = generate_codename(guardian_name+name, requiredlen)
        count = 1
        codename = ''
        # codename = generate_codename(data[gi], whichdepartment)
        codename = generate_codename(data[gi])
        while (f'{codename}{count:03}' in created_codenames):
            count += 1
        codename = f'{codename}{count:03}'
        created_codenames.add(codename)

        pdfile = FPDF()
        pdfile.set_margins(0, 0, 0)
        pdfile.set_font('Helvetica')
        # pdfile.set_font('Arial')
        # pdfile.set_font('Lucida Sans Unicode')
        # pdfile.set_font('Symbol')
        for i, imgfile in enumerate(__IMGS, start=1):
            pdfile.add_page()
            pdfile.image(f'{__IMGDIR}/{imgfile}', 0, 0, w=__PG_WIDTH)
            # Page 1
            if i == 1:
                stname = data[gi][0]
                pdfile.ln(230)
                pdfile.set_x(8)
                pdfile.set_font_size(25)
                # pdfile.cell(100, 10, codename, border=0)
                pdfile.cell(100, 10, stname, border=0)
                pdfile.set_font_size(__DEF_FSIZE)
            # Page 4
            elif i == 4:
                # categories      = ['Variables', 'Anxiety (AN)','Depression (DP)','Suicidal Ideation (SI)',
                #     'Substance Abuse (SA)', 'Self-esteem Problems (SE)', 'Interpersonal Problems (IP)',
                #     'Family Problems (FP)', 'Academic Problems (AP)','Career Problems (CP)']
                
                plot_tscores(pdfile, reco.CATEGORIES, tscores, problemareas)
                pdfile.image(f'{__GRAPHS_DIR}/{graphs[stu]}', 18, 128, h=120)
            # Page 5
            elif i == 5:
                pdfile.set_margins(20, 0, 20)
                pdfile.set_font_size(__DEF_FSIZE+5)
                pdfile.set_y(35)
                # for ii, keys in enumerate(problemareas[1:]):
                for ii, keys in enumerate(problemareas):
                    for j, recommend in enumerate(reco.crpair[keys]):
                        # recommend = "` " + recommend
                        # recommend = "~ " + recommend
                        advice = recommend.encode('utf-8').decode('iso-8859-1')
                        # pdfile.write(txt="SAMPLE TEXT SAMPLE TEXT SAMPLETEAKJLDSFLKAJDSF; JADLFKJALKS;JFL;AKJDFLKAJ;FKLAJDSFKL;JADLSF")

                        # pdfile.set_x(20)
                        if (pdfile.get_y() > 165):
                            pdfile.ln(210)
                            pdfile.set_x(40)
                            # pdfile.cell(70, 20, "sample", border=1)
                            pdfile.link(40, 210, 70, 20, ACROCARE_EMAIL)
                            pdfile.add_page()
                            pdfile.image(f'{__IMGDIR}/{TEMPLATE_WITH_CONTACT}', 0, 0, w=__PG_WIDTH)
                            pdfile.set_y(35)
                        pdfile.set_font(style="B")
                        # pdfile.write(txt="\n" + advice)
                        pdfile.write(txt="\n~ ") 
                        # pdfile.write(txt="\n. ")
                        # pdfile.write(txt="\no ")
                        # pdfile.write(txt="\nO ")
                        pdfile.set_font(style="")
                        pdfile.write(txt=advice)
                        # pdfile.write(txt="\n\u2022 " + advice)
                        # pdfile.write(h=__DEF_FSIZE, txt=advice)
                        # pdfile.write(h=8, txt="\n"+advice)
                        pdfile.ln()
                        if debug_setting:
                            print(advice)
                            print(pdfile.get_y())

                pdfile.set_font(style="B")
                pdfile.ln(10)
                pdfile.write(txt="Do not hesitate to connect with counselor to seek professional help immediately.")
                pdfile.set_font(style="")
                
                pdfile.ln(210)
                pdfile.set_x(40)
                # pdfile.cell(70, 20, "sample", border=1)
                pdfile.link(40, 210, 70, 20, ACROCARE_EMAIL)


            # Page 6
            elif i == 6:
                pdfile.ln(50)
                pdfile.link(0, 50, __PG_WIDTH, 130, AITRGOOGLE_URL)

        # outfile_name = f"{gi}-{stu}.pdf"
        # pdfile.output(f"{dirpath}/{outfile_name}")
        outfile_name = f"{gi}-{codename}.pdf"
        pdfile.output(os.path.join(dirpath,outfile_name))
        
        con1.commit()
        # SURVEY_ID, STU_NAME, EMAIL, GUARDIAN, AGE, GENDER, MOBILE, INSTITUTE, STREAM, YEAR, MONTHLY_FAMILY_INCOME, FILE, CODENAME
        cur.execute("""INSERT INTO tblSurveyReports(survey_id, student_codename, 
            student_name, email, guardian_name, age, gender, mobile_number, 
            institute, stream, year, monthly_family_income, tot_problem_areas,
            problem_areas, report_file) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (survey_id, codename, *data[gi], len(problemareas),', '.join(problemareas), outfile_name))
        con1.commit()
    con1.close()
    clean_up(__GRAPHS_DIR)

    return dirpath
