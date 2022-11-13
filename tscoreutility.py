import config
import csv
import os
import openpyxl as op
import matplotlib.pyplot as plt
from dirutility import create_dir

def Parse_Csv_To_List(path):
    
    with open(path) as file_object:
        reader_file = csv.reader(file_object)
        Raw_Data = list(reader_file)
    return Raw_Data


def Parse_Excel_To_List(path):
    """ 
        Parse a given excel sheet and return list. 
        The data to be parsed should be either in first sheet of excel file or 
        sheet should be named 'Raw Data' (case insensitive).
    """
    WorkBook = op.load_workbook(path)       #WorkBook in which data is available
    DEFAULT_SHEET = "Raw Data"
    try:
        Sheet = WorkBook[DEFAULT_SHEET]          #Particular Sheet in Workbook
    except KeyError as kerr:
        # There is a possiblity that excel file contains 'Raw data' sheet but
        # the letters could be in uppercase or lowercase and cause a mismatch
        DEFAULT_SHEET = DEFAULT_SHEET.lower()
        for sheet in WorkBook.sheetnames:
            if DEFAULT_SHEET == sheet.lower():
                Sheet = sheet
                break
        else:
        # if for-loop doesn't break then it means there was no Raw data sheet in
        # excel file and thus as backup we select the first sheet
            FIRST_SHEET = WorkBook.sheetnames[0]
            Sheet = WorkBook[FIRST_SHEET]
            # Sheet = WorkBook.active               #Last sheet found in excel        
    except Exception as e:
        # Unknown/Unexpected error occurred.
        print(f"Exception, at line {e.__traceback__.tb_lineno},",  *e.args)
        print(e.__traceback__.tb_frame, "\n")
        return []

    Total_rows = Sheet.max_row              #Total no. of rows in Sheet
    Total_clms = Sheet.max_column           #Total no. of coloms in Sheet

    RawData = []

    for i in range(1,Total_rows+1):         #Sheet data --->Data(It is 2d list)
        temp = []
        for j in range(1,Total_clms+1):
            temp.append(str(Sheet.cell(i,j).value))
        RawData.append(temp)

    WorkBook.close()
    return RawData


def Scores(value,clmNo):
    if clmNo==20 or clmNo==29 or clmNo==33 or clmNo==38 or clmNo==47 or clmNo==52 or clmNo==58 or clmNo==65 or clmNo==84 or clmNo==101:
       d={
        #'FALSE':4,
        'false':4,
        'slightly true':3,
        'mainly true':2,
        'very true':1
         } 
    else:
        d={
        #'FALSE':1,
        'false':1,
        'slightly true':2,
        'mainly true':3,
        'very true':4
         }     
    return d.get(value.strip().lower(), value)


def T_Scores(Name,value):
    d = {"AN": {12: 30, 13: 34, 14: 38, 15: 40, 16: 43, 17: 44, 18: 46, 19: 48, 20: 49, 21: 51, 22: 52, 23: 53, 24: 54, 25: 56, 26: 57, 27: 58, 28: 59, 29: 60, 30: 61, 31: 62, 32: 63, 33: 64, 34: 65, 35: 66, 36: 67, 37: 68, 38: 69, 39: 70, 40: 71, 41: 72, 42: 73, 43: 74, 44: 75, 45: 77, 46: 78, 47: 78, 48: 80},
         "DP": {12: 27, 13: 38, 14: 43, 15: 46, 16: 49, 17: 51, 18: 53, 19: 54, 20: 55, 21: 57, 22: 58, 23: 59, 24: 60, 25: 61, 26: 62, 27: 63, 28: 64, 29: 65, 30: 66, 31: 67, 32: 68, 33: 69, 34: 70, 35: 71, 36: 72, 37: 73, 38: 74, 39: 75, 40: 77, 41: 78, 42: 78, 43: 78, 44: 79, 45: 79, 46: 79, 47: 80, 48: 80},
         "SI": {12: 44, 13: 50, 14: 54, 15: 56, 16: 59, 17: 60, 18: 62, 19: 63, 20: 64, 21: 65, 22: 66, 23: 67, 24: 68, 25: 69, 26: 69, 27: 70, 28: 71, 29: 71, 30: 72, 31: 73, 32: 73, 33: 73, 34: 74, 35: 74, 36: 74, 37: 75, 38: 75, 39: 75, 40: 76, 41: 76, 42: 76, 43: 77, 44: 77, 45: 78, 46: 78, 47: 79, 48: 80},
         "SA": {12: 39, 13: 45, 14: 48, 15: 51, 16: 53, 17: 54, 18: 56, 19: 57, 20: 58, 21: 59, 22: 60, 23: 61, 24: 62, 25: 63, 26: 64, 27: 65, 28: 65, 29: 66, 30: 67, 31: 68, 32: 69, 33: 70, 34: 70, 35: 71, 36: 72, 37: 73, 38: 74, 39: 75, 40: 75, 41: 76, 42: 76, 43: 77, 44: 77, 45: 78, 46: 78, 47: 79, 48: 80},
         "SE": {12: 23, 13: 30, 14: 34, 15: 38, 16: 40, 17: 43, 18: 45, 19: 47, 20: 48, 21: 50, 22: 51, 23: 53, 24: 54, 25: 55, 26: 57, 27: 58, 28: 59, 29: 60, 30: 61, 31: 62, 32: 63, 33: 64, 34: 65, 35: 66, 36: 67, 37: 68, 38: 69, 39: 70, 40: 71, 41: 72, 42: 73, 43: 74, 44: 75, 45: 76, 46: 77, 47: 78, 48: 80},
         "IP": {12: 29, 13: 34, 14: 37, 15: 40, 16: 43, 17: 45, 18: 47, 19: 48, 20: 50, 21: 52, 22: 53, 23: 54, 24: 56, 25: 57, 26: 58, 27: 59, 28: 60, 29: 62, 30: 63, 31: 64, 32: 65, 33: 66, 34: 67, 35: 68, 36: 70, 37: 71, 38: 72, 39: 73, 40: 74, 41: 75, 42: 77, 43: 78, 44: 80, 45: 80, 46: 80, 47: 80, 48: 80},
         "FP": {12: 32, 13: 38, 14: 42, 15: 45, 16: 47, 17: 49, 18: 51, 19: 52, 20: 54, 21: 55, 22: 56, 23: 57, 24: 58, 25: 59, 26: 60, 27: 61, 28: 62, 29: 63, 30: 64, 31: 65, 32: 66, 33: 66, 34: 67, 35: 68, 36: 69, 37: 70, 38: 71, 39: 71, 40: 72, 41: 73, 42: 74, 43: 75, 44: 76, 45: 77, 46: 78, 47: 79, 48: 80},
         "AP": {12: 28, 13: 31, 14: 34, 15: 36, 16: 38, 17: 40, 18: 42, 19: 44, 20: 45, 21: 47, 22: 48, 23: 50, 24: 51, 25: 52, 26: 54, 27: 55, 28: 56, 29: 57, 30: 58, 31: 60, 32: 61, 33: 62, 34: 63, 35: 64, 36: 65, 37: 66, 38: 67, 39: 69, 40: 70, 41: 71, 42: 72, 43: 73, 44: 74, 45: 75, 46: 76, 47: 77, 48: 78},
         "CP": {12: 36, 13: 41, 14: 44, 15: 46, 16: 48, 17: 50, 18: 51, 19: 52, 20: 53, 21: 54, 22: 55, 23: 56, 24: 57, 25: 58, 26: 59, 27: 59, 28: 60, 29: 61, 30: 62, 31: 62, 32: 63, 33: 64, 34: 65, 35: 65, 36: 66, 37: 67, 38: 68, 39: 69, 40: 69, 41: 70, 42: 71, 43: 72, 44: 74, 45: 75, 46: 76, 47: 78, 48: 80}
         }
    return d[Name][value]


def Tscores_And_Students_ImpData(RawData):
    Total_rows = len(RawData)      #Total no. of rows in Data
    Total_cols = len(RawData[0])   #Total no. of coloumns in Data
    # print(Total_rows, Total_cols)

    for i in range(0,Total_rows):
        for j in range(0,Total_cols):
            RawData[i][j] = Scores(RawData[i][j],j)     #Findinf the values of False,Very True...

    Student_Imp_Data = []
    AllQuestions = []
    Raw_Scores = [['AN','DP','SI','SA','SE','IP','FP','AP','CP']]
    for i in range(0,Total_rows):
        student = []                                #Storing Students data
        OneRowQuestions = []                        # Storing rows questions
        for j in range(3,Total_cols):
            if j<13:
                student.append(RawData[i][j])
            else:
                OneRowQuestions.append(RawData[i][j])    
        Student_Imp_Data.append(student)
        AllQuestions.append(OneRowQuestions)

    T_Scores_Of_All_Students = [['T-AN','T-DP','T-SI','T-SA','T-SE','T-IP','T-FP','T-AP','T-CP']]
    for i in range(1,Total_rows):
        """
        This code could be simplified using list comprehension and sum() function 
            AN = sum( [RawData[i][num] for num in range(14, 114, 9)] )
            DP = sum( [RawData[i][num] for num in range(16, 116, 9)] )
            SI = sum( ... )
            ...
        """
        temp = []
        temp_Raw_Scores = []
        count = 1
        j = 13
        AN = DP = SI = SA = SE = IP = FP = AP = CP = 0
        while count <= 12:
            AP += RawData[i][j]
            AN += RawData[i][j+1]
            IP += RawData[i][j+2]
            DP += RawData[i][j+3]
            CP += RawData[i][j+4]
            SI += RawData[i][j+5]
            SA += RawData[i][j+6]
            SE += RawData[i][j+7]
            FP += RawData[i][j+8]
            j += 9
            count += 1
        temp.append(T_Scores("AN", AN))
        temp.append(T_Scores("DP", DP))
        temp.append(T_Scores("SI", SI))
        temp.append(T_Scores("SA", SA))
        temp.append(T_Scores("SE", SE))
        temp.append(T_Scores("IP", IP))
        temp.append(T_Scores("FP", FP))
        temp.append(T_Scores("AP", AP))
        temp.append(T_Scores("CP", CP))
        T_Scores_Of_All_Students.append(temp)

        temp_Raw_Scores.append(AN)
        temp_Raw_Scores.append(DP)
        temp_Raw_Scores.append(SI)
        temp_Raw_Scores.append(SA)
        temp_Raw_Scores.append(SE)
        temp_Raw_Scores.append(IP)
        temp_Raw_Scores.append(FP)
        temp_Raw_Scores.append(AP)
        temp_Raw_Scores.append(CP)
        Raw_Scores.append(temp_Raw_Scores)

    return T_Scores_Of_All_Students, Student_Imp_Data, AllQuestions, Raw_Scores


def Graph_Of_Tscores(Y_Axis_tscore, Student_Data, i, directory_path):
    
    X_Axis_Names = ['Anxiety','Depression','Suicidal Ideation','Substance Abuse','Self-esteem Problems','Interpersonal Problems','Family Problems','Academic Problems','Career Problems']

    plt.rcParams['figure.figsize']=(8,5)        #Size(Area) of Graph
    plt.figure(facecolor="#BFEFFF")             #Background Color

    plt.plot(X_Axis_Names,Y_Axis_tscore,color="red",marker='o') #Plot graph
    plt.ylim(0,80)
    plt.xticks(rotation=50,ha="center")
    plt.gcf().subplots_adjust(bottom=0.286)     #Adjusting(cutting) the X-Axis names
    plt.grid(alpha=0.7)                         #For grid

    for k in range(len(X_Axis_Names)):
        plt.text(k,Y_Axis_tscore[k],Y_Axis_tscore[k],ha='center',va="bottom")   #For Adding value in line 
        
    Graph_Name = Student_Data[i][0] + str(i) + ".png"
    # plt.savefig(directory_path + '/' + Student_Data[i][0] + str(i) + ".png",bbox_inches="tight")
    plt.savefig(os.path.join(directory_path, Graph_Name), bbox_inches="tight")

    return Graph_Name


def Plot_Graphs(TScores, Student_Data):
    Total_rows = len(TScores)      #Total no. of rows in Data
    Total_cols = len(TScores[0])   #Total no. of coloms in Data

    # Instead of list, a mapping of student-name to graph-image-path is more convenient and secure.
    # List_Of_Graphs = []
    folder = create_dir(config.TEMPDIR)
    Graphs = {}
    for i in range(1,Total_rows):
        Y_Axis_tscore = []
        for j in range(0,Total_cols):
            Y_Axis_tscore.append(TScores[i][j])
        # List_Of_Graphs.append(Graph_Of_Tscores(Y_Axis_tscore,Student_Data,i))  #For Making graph of row student
        
        Graphs[Student_Data[i][0]] = Graph_Of_Tscores(Y_Axis_tscore,Student_Data,i,folder)
    
    # return List_Of_Graphs
    
    return Graphs
    # return {'abcd': 'sample-plot.png'}


if __name__ == '__main__':
    from pprint import pprint
    FILE = 'cas-copy.xlsx'
    FILE = '1-record-cas-sheet.xlsx'
    data = Parse_Excel_To_List(FILE)
    # pprint(data)
    tscdata, studata, allquestions, rawscores = Tscores_And_Students_ImpData(data)
    pprint(studata)
    pprint(tscdata)
    # pprint(rawscores)